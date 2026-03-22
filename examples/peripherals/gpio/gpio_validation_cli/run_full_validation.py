#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shlex
import subprocess
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import serial
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PASS_FILL = PatternFill(fill_type="solid", fgColor="00B050")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
PASS_FONT = Font(bold=True, color="FFFFFF")
FAIL_FONT = Font(bold=True, color="9C0006")
HEADER_FONT = Font(bold=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)

GPIO_OK_PREFIX = "GPIOV OK"
GPIO_FAIL_PREFIX = "GPIOV FAIL"
TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"
ANSI_ESCAPE_RE = re.compile(r"\x1b\[[0-9;?]*[ -/]*[@-~]")
LEVEL_INTERRUPT_TRIGGERS = {"sync_low", "sync_high", "async_low", "async_high"}
DEFAULT_IRQ_GET_MANY_CHUNK_SIZE = 8
DEFAULT_IRQ_READ_RETRY_MAX = 2
DEFAULT_IRQ_READ_RETRY_DELAY_MS = 20


class ValidationRuntimeError(RuntimeError):
    pass


@dataclass(frozen=True)
class SerialConfig:
    label: str
    port: str
    baudrate: int
    run_reset_on_open: bool
    reset_pulse_ms: int
    boot_wait_ms: int
    dtr_low_is_true: bool
    rts_low_is_true: bool


@dataclass(frozen=True)
class PinGroup:
    name: str
    dut_pins: list[int]
    tester_pins: list[int]


@dataclass(frozen=True)
class CommandResult:
    ok: bool
    line: str
    fields: dict[str, str]
    elapsed_ms: int


class ArtifactLogger:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._fp = self.path.open("w", encoding="utf-8")
        self._lock = threading.Lock()

    def log(self, channel: str, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
        with self._lock:
            self._fp.write(f"[{timestamp}] [{channel}] {message}\n")
            self._fp.flush()

    def close(self) -> None:
        with self._lock:
            self._fp.close()


class SerialEndpoint:
    def __init__(self, config: SerialConfig, logger: ArtifactLogger, startup_drain_sec: float) -> None:
        self.config = config
        self.logger = logger
        self.startup_drain_sec = startup_drain_sec
        self._serial: serial.Serial | None = None
        self._lock = threading.Lock()
        self._buffer = ""

    def __enter__(self) -> "SerialEndpoint":
        self.open()
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def open(self) -> None:
        if self._serial is not None:
            return

        self.logger.log(self.config.label, f"OPEN port={self.config.port} baudrate={self.config.baudrate}")
        self._serial = serial.Serial(
            port=self.config.port,
            baudrate=self.config.baudrate,
            timeout=0.1,
            write_timeout=1,
            exclusive=True,
        )
        if self.config.run_reset_on_open:
            self.reset_to_run_mode()
        self.drain(self.startup_drain_sec)

    def close(self) -> None:
        if self._serial is None:
            return

        self.logger.log(self.config.label, "CLOSE")
        self._serial.close()
        self._serial = None

    def ping(self, retries: int = 3, timeout_sec: float = 2.0) -> CommandResult:
        last_error: Exception | None = None
        for attempt in range(1, retries + 1):
            try:
                self.logger.log(self.config.label, f"PING attempt={attempt}")
                result = self.send_command("gpiov_ping", timeout_sec)
                if not result.ok:
                    raise ValidationRuntimeError(result.line)
                return result
            except Exception as exc:  # pragma: no cover - hardware path
                last_error = exc
                self.logger.log(self.config.label, f"PING_RETRY attempt={attempt} error={exc}")
                time.sleep(0.2)

        raise ValidationRuntimeError(
            f"{self.config.label} did not respond to gpiov_ping on {self.config.port}: {last_error}"
        )

    def reset_to_run_mode(self) -> None:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        self.logger.log(
            self.config.label,
            "RESET run_mode"
            + f" dtr_low_is_true={int(self.config.dtr_low_is_true)}"
            + f" rts_low_is_true={int(self.config.rts_low_is_true)}"
            + f" reset_pulse_ms={self.config.reset_pulse_ms}"
            + f" boot_wait_ms={self.config.boot_wait_ms}",
        )
        self._set_control_lines(dtr_low=True, rts_low=True)
        time.sleep(self.config.reset_pulse_ms / 1000.0)
        self._set_control_lines(dtr_low=True, rts_low=False)
        time.sleep(self.config.boot_wait_ms / 1000.0)
        self._serial.reset_input_buffer()
        self._serial.reset_output_buffer()

    def drain(self, duration_sec: float) -> None:
        if self._serial is None or duration_sec <= 0:
            return

        deadline = time.monotonic() + duration_sec
        while True:
            line = self._read_next_line(deadline)
            if line is None:
                break
            text = line.strip()
            if text:
                self.logger.log(self.config.label, f"RX {text}")

    def send_command(self, command: str, timeout_sec: float) -> CommandResult:
        if self._serial is None:
            raise ValidationRuntimeError(f"{self.config.label} is not open")

        with self._lock:
            self.drain(0.05)
            start = time.monotonic()
            payload = f"{command}\r\n".encode("utf-8")
            self.logger.log(self.config.label, f"TX {command}")
            self._serial.write(payload)
            self._serial.flush()

            deadline = start + timeout_sec
            while True:
                line = self._read_next_line(deadline)
                if line is None:
                    break

                text = line.strip()
                if not text:
                    continue

                self.logger.log(self.config.label, f"RX {text}")
                clean_text = strip_ansi(text)
                if clean_text.startswith(GPIO_OK_PREFIX) or clean_text.startswith(GPIO_FAIL_PREFIX):
                    return CommandResult(
                        ok=clean_text.startswith(GPIO_OK_PREFIX),
                        line=clean_text,
                        fields=parse_key_value_fields(clean_text),
                        elapsed_ms=int((time.monotonic() - start) * 1000),
                    )

        raise ValidationRuntimeError(
            f"{self.config.label} command timeout after {timeout_sec:.2f}s: {command}"
        )

    def _read_next_line(self, deadline: float) -> str | None:
        if self._serial is None:
            return None

        while time.monotonic() < deadline:
            if "\n" in self._buffer:
                line, self._buffer = self._buffer.split("\n", 1)
                return line.rstrip("\r")

            waiting = self._serial.in_waiting
            chunk = self._serial.read(waiting or 1)
            if not chunk:
                continue

            self._buffer += chunk.decode("utf-8", errors="replace")

        return None

    def _set_control_lines(self, *, dtr_low: bool, rts_low: bool) -> None:
        if self._serial is None:
            return

        self._serial.dtr = dtr_low if self.config.dtr_low_is_true else not dtr_low
        self._serial.rts = rts_low if self.config.rts_low_is_true else not rts_low


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="One-click GPIO validation entrypoint: DUT build + DUT flash + tester/DUT serial validation"
    )
    parser.add_argument(
        "--config",
        default="gpio_validation_config.example.json",
        help="Unified GPIO validation config JSON. Default: gpio_validation_config.example.json",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Directory for logs, report, and result JSON. Default: validation_runs/<timestamp>",
    )
    parser.add_argument(
        "--skip-build",
        action="store_true",
        help="Skip DUT build even if enabled in the config",
    )
    parser.add_argument(
        "--skip-flash",
        action="store_true",
        help="Skip DUT flash even if enabled in the config",
    )
    parser.add_argument(
        "--skip-test",
        action="store_true",
        help="Skip the GPIO serial validation stage",
    )
    return parser.parse_args()


def load_json(path: Path) -> dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise SystemExit(f"Config file not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Invalid JSON in config file {path}: {exc}") from exc

    if not isinstance(data, dict):
        raise SystemExit("Config root must be a JSON object")
    return data


def ensure_dict(raw: Any, field_name: str) -> dict[str, Any]:
    if raw is None:
        return {}
    if not isinstance(raw, dict):
        raise SystemExit(f"{field_name} must be a JSON object")
    return raw


def ensure_bool(raw: Any, field_name: str, default: bool) -> bool:
    if raw is None:
        return default
    if not isinstance(raw, bool):
        raise SystemExit(f"{field_name} must be a boolean")
    return raw


def ensure_str(raw: Any, field_name: str, allow_empty: bool = False) -> str:
    if not isinstance(raw, str):
        raise SystemExit(f"{field_name} must be a string")
    if (not allow_empty) and (not raw):
        raise SystemExit(f"{field_name} must be a non-empty string")
    return raw


def ensure_str_list(raw: Any, field_name: str) -> list[str]:
    if raw is None:
        return []
    if not isinstance(raw, list) or not all(isinstance(item, str) for item in raw):
        raise SystemExit(f"{field_name} must be an array of strings")
    return list(raw)


def parse_serial_config(raw: Any, label: str) -> SerialConfig:
    if not isinstance(raw, dict):
        raise SystemExit(f"serial.{label} must be a JSON object")

    port = raw.get("port")
    baudrate = raw.get("baudrate", 2_000_000)
    run_reset_on_open = raw.get("run_reset_on_open", True)
    reset_pulse_ms = raw.get("reset_pulse_ms", 5)
    boot_wait_ms = raw.get("boot_wait_ms", 500)
    dtr_low_is_true = raw.get("dtr_low_is_true", True)
    rts_low_is_true = raw.get("rts_low_is_true", True)

    if not isinstance(port, str) or not port:
        raise SystemExit(f"serial.{label}.port must be a non-empty string")
    if not isinstance(baudrate, int) or baudrate <= 0:
        raise SystemExit(f"serial.{label}.baudrate must be a positive integer")
    if not isinstance(run_reset_on_open, bool):
        raise SystemExit(f"serial.{label}.run_reset_on_open must be a boolean")
    if not isinstance(reset_pulse_ms, int) or reset_pulse_ms <= 0:
        raise SystemExit(f"serial.{label}.reset_pulse_ms must be a positive integer")
    if not isinstance(boot_wait_ms, int) or boot_wait_ms < 0:
        raise SystemExit(f"serial.{label}.boot_wait_ms must be a non-negative integer")
    if not isinstance(dtr_low_is_true, bool):
        raise SystemExit(f"serial.{label}.dtr_low_is_true must be a boolean")
    if not isinstance(rts_low_is_true, bool):
        raise SystemExit(f"serial.{label}.rts_low_is_true must be a boolean")

    return SerialConfig(
        label=label.upper(),
        port=port,
        baudrate=baudrate,
        run_reset_on_open=run_reset_on_open,
        reset_pulse_ms=reset_pulse_ms,
        boot_wait_ms=boot_wait_ms,
        dtr_low_is_true=dtr_low_is_true,
        rts_low_is_true=rts_low_is_true,
    )


def parse_pin_groups(raw_groups: Any) -> list[PinGroup]:
    if not isinstance(raw_groups, list) or not raw_groups:
        raise SystemExit("pin_groups must be a non-empty array")

    groups: list[PinGroup] = []
    for index, item in enumerate(raw_groups, start=1):
        if not isinstance(item, dict):
            raise SystemExit(f"pin_groups[{index}] must be an object")

        name = item.get("name", f"group_{index}")
        dut_pins = parse_pin_list(item.get("dut_pins"), f"pin_groups[{index}].dut_pins")
        tester_pins = parse_pin_list(item.get("tester_pins"), f"pin_groups[{index}].tester_pins")
        if len(dut_pins) != len(tester_pins):
            raise SystemExit(f"pin_groups[{index}] dut_pins and tester_pins must have the same length")

        groups.append(PinGroup(name=str(name), dut_pins=dut_pins, tester_pins=tester_pins))

    return groups


def parse_pin_list(raw: Any, field_name: str) -> list[int]:
    if not isinstance(raw, list) or not raw:
        raise SystemExit(f"{field_name} must be a non-empty integer array")

    pins: list[int] = []
    seen: set[int] = set()
    for value in raw:
        if not isinstance(value, int) or value < 0:
            raise SystemExit(f"{field_name} must contain only non-negative integers")
        if value in seen:
            raise SystemExit(f"{field_name} contains duplicate pin {value}")
        pins.append(value)
        seen.add(value)

    return pins


def normalize_pull_modes(raw: Any) -> list[str]:
    if raw is None:
        return ["float"]
    if isinstance(raw, str):
        return [normalize_pull_name(raw)]
    if isinstance(raw, list) and raw:
        return [normalize_pull_name(item) for item in raw]
    raise SystemExit("dut_pull must be a string or a non-empty string array")


def normalize_pull_name(name: Any) -> str:
    if not isinstance(name, str):
        raise SystemExit("pull value must be a string")

    normalized = name.strip().lower()
    aliases = {
        "float": "float",
        "up": "up",
        "down": "down",
        "pull_up": "up",
        "pull_down": "down",
        "pullup": "up",
        "pulldown": "down",
    }
    if normalized not in aliases:
        raise SystemExit(f"Unsupported pull mode: {name}")
    return aliases[normalized]


def parse_interrupt_triggers(raw_interrupt: dict[str, Any]) -> list[dict[str, Any]]:
    triggers = raw_interrupt.get("triggers")
    if triggers is None and "trigger" in raw_interrupt:
        triggers = [
            {
                "trigger": raw_interrupt.get("trigger"),
                "idle_level": raw_interrupt.get("idle_level"),
                "active_level": raw_interrupt.get("active_level"),
                "expected_delta": raw_interrupt.get("expected_delta"),
            }
        ]

    if not isinstance(triggers, list) or not triggers:
        raise SystemExit("scenarios.interrupt.triggers must be a non-empty array when interrupt is enabled")

    parsed: list[dict[str, Any]] = []
    for index, item in enumerate(triggers, start=1):
        if not isinstance(item, dict):
            raise SystemExit(f"scenarios.interrupt.triggers[{index}] must be an object")

        trigger = item.get("trigger")
        idle_level = item.get("idle_level")
        active_level = item.get("active_level")
        expected_delta = item.get("expected_delta")
        timeout_ms = item.get("timeout_ms")
        active_hold_ms = item.get("active_hold_ms")
        name = item.get("name", trigger)

        if not isinstance(trigger, str) or not trigger:
            raise SystemExit(f"scenarios.interrupt.triggers[{index}].trigger must be a non-empty string")
        if idle_level not in (0, 1):
            raise SystemExit(f"scenarios.interrupt.triggers[{index}].idle_level must be 0 or 1")
        if active_level not in (0, 1):
            raise SystemExit(f"scenarios.interrupt.triggers[{index}].active_level must be 0 or 1")
        if expected_delta is not None and (not isinstance(expected_delta, int) or expected_delta <= 0):
            raise SystemExit(f"scenarios.interrupt.triggers[{index}].expected_delta must be a positive integer")
        if timeout_ms is not None and (not isinstance(timeout_ms, int) or timeout_ms <= 0):
            raise SystemExit(f"scenarios.interrupt.triggers[{index}].timeout_ms must be a positive integer")
        if active_hold_ms is not None and (not isinstance(active_hold_ms, int) or active_hold_ms <= 0):
            raise SystemExit(f"scenarios.interrupt.triggers[{index}].active_hold_ms must be a positive integer")

        parsed.append(
            {
                "name": str(name),
                "trigger": trigger,
                "idle_level": idle_level,
                "active_level": active_level,
                "expected_delta": expected_delta,
                "timeout_ms": timeout_ms,
                "active_hold_ms": active_hold_ms,
            }
        )

    return parsed


def parse_key_value_fields(line: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    for token in line.split()[2:]:
        if "=" not in token:
            continue
        key, value = token.split("=", 1)
        fields[key] = value
    return fields


def strip_ansi(text: str) -> str:
    return ANSI_ESCAPE_RE.sub("", text).strip()


def pins_to_csv(pins: list[int]) -> str:
    return ",".join(str(pin) for pin in pins)


def chunk_pins(pins: list[int], chunk_size: int) -> list[list[int]]:
    return [pins[index : index + chunk_size] for index in range(0, len(pins), chunk_size)]


def bool_enabled(raw: Any, default: bool = True) -> bool:
    if raw is None:
        return default
    if isinstance(raw, bool):
        return raw
    raise SystemExit("Scenario enabled field must be a boolean")


def scenario_timeout_ms(raw: dict[str, Any], field_name: str, default: int) -> int:
    value = raw.get(field_name, default)
    if not isinstance(value, int) or value <= 0:
        raise SystemExit(f"{field_name} must be a positive integer")
    return value


def scenario_non_negative_int(raw: dict[str, Any], field_name: str, default: int) -> int:
    value = raw.get(field_name, default)
    if not isinstance(value, int) or value < 0:
        raise SystemExit(f"{field_name} must be a non-negative integer")
    return value


def result_excerpt(result: CommandResult) -> str:
    for key in ("levels", "deltas", "irq_counts"):
        value = result.fields.get(key)
        if value:
            return value
    return result.line


def is_level_interrupt_trigger(trigger: str) -> bool:
    return trigger in LEVEL_INTERRUPT_TRIGGERS


def parse_pin_value_map(raw: str, field_name: str) -> dict[int, int]:
    if not raw:
        raise ValueError(f"missing {field_name}")

    parsed: dict[int, int] = {}
    for item in raw.split(","):
        token = item.strip()
        if not token:
            continue
        if ":" not in token:
            raise ValueError(f"invalid {field_name} entry: {token}")
        pin_text, value_text = token.split(":", 1)
        try:
            pin = int(pin_text)
            value = int(value_text)
        except ValueError as exc:
            raise ValueError(f"invalid {field_name} entry: {token}") from exc
        parsed[pin] = value

    if not parsed:
        raise ValueError(f"empty {field_name}")
    return parsed


def evaluate_level_interrupt_result(
    group: PinGroup,
    irq_counts: dict[int, int],
    target_pin: int,
    expected_delta: int,
) -> tuple[str, list[str]]:
    response_text = ",".join(f"{pin}:{irq_counts.get(pin, '?')}" for pin in group.dut_pins)
    detail_parts: list[str] = []

    target_delta = irq_counts.get(target_pin)
    if target_delta is None:
        detail_parts.append(f"missing target irq count pin={target_pin} irq_counts={response_text}")
        return response_text, detail_parts
    if target_delta < expected_delta:
        detail_parts.append(
            f"target_irq_too_low target_pin={target_pin} expected_min_delta={expected_delta} "
            f"actual_delta={target_delta} irq_counts={response_text}"
        )

    for pin in group.dut_pins:
        delta = irq_counts.get(pin)
        if delta is None:
            detail_parts.append(f"missing irq count pin={pin} irq_counts={response_text}")
            return response_text, detail_parts
        if pin != target_pin and delta != 0:
            detail_parts.append(f"unexpected_irq pin={pin} target_pin={target_pin} irq_counts={response_text}")
            return response_text, detail_parts

    return response_text, detail_parts


def parse_irq_counts_chunk(result: CommandResult, chunk: list[int], chunk_csv: str) -> tuple[dict[int, int] | None, str | None]:
    if not result.ok:
        return None, result.line

    raw_irq_counts = result.fields.get("irq_counts", "")
    try:
        parsed = parse_pin_value_map(raw_irq_counts, "irq_counts")
    except ValueError as exc:
        return None, f"{exc} chunk={chunk_csv}"

    values: dict[int, int] = {}
    for pin in chunk:
        value = parsed.get(pin)
        if value is None:
            return None, f"missing irq count pin={pin} chunk={chunk_csv} irq_counts={raw_irq_counts}"
        values[pin] = value
    return values, None


def read_irq_counts_chunked(
    dut: SerialEndpoint,
    pins: list[int],
    *,
    chunk_size: int,
    retry_max: int,
    retry_delay_sec: float,
    log_fp,
    log_prefix: str,
) -> tuple[dict[int, int], list[str], int]:
    combined: dict[int, int] = {}
    errors: list[str] = []
    retry_count = 0
    for chunk in chunk_pins(pins, chunk_size):
        chunk_csv = pins_to_csv(chunk)
        last_error = ""
        for attempt in range(retry_max + 1):
            if attempt > 0:
                retry_count += 1
                log_progress(
                    log_fp,
                    "interrupt",
                    f"{log_prefix} irq_read_retry chunk={chunk_csv} retry={attempt}/{retry_max} reason={shorten_text(last_error)}",
                )
                if retry_delay_sec > 0:
                    time.sleep(retry_delay_sec)

            try:
                result = dut.send_command(f"gpiov_irq_get_many {chunk_csv}", 3.0)
            except Exception as exc:  # pragma: no cover - hardware path
                last_error = f"{exc} chunk={chunk_csv}"
            else:
                parsed_chunk, parse_error = parse_irq_counts_chunk(result, chunk, chunk_csv)
                if parse_error is None:
                    combined.update(parsed_chunk or {})
                    if attempt > 0:
                        log_progress(
                            log_fp,
                            "interrupt",
                            f"{log_prefix} irq_read_retry chunk={chunk_csv} recovered retry_count={attempt}",
                        )
                    last_error = ""
                    break
                last_error = parse_error

        if last_error:
            if retry_max > 0:
                last_error = f"{last_error} retry_count={retry_max}"
            errors.append(last_error)

    return combined, errors, retry_count


def shorten_text(text: str, limit: int = 160) -> str:
    normalized = strip_ansi(str(text)).replace("\r", " ").replace("\n", " ")
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if len(normalized) <= limit:
        return normalized
    return normalized[: limit - 3] + "..."


def log_progress(log_fp, stage: str, message: str) -> None:
    timestamp = datetime.now().strftime("%H:%M:%S.%f")[:-3]
    line = f"[{timestamp}] [{stage}] {message}"
    print(line)
    log_fp.write(line + "\n")
    log_fp.flush()


def log_stage_summary(log_fp, stage: str, records: list[dict[str, Any]]) -> None:
    passed = sum(1 for record in records if record["result"] == "PASS")
    failed = len(records) - passed
    log_progress(log_fp, stage, f"summary pass={passed} fail={failed} total={len(records)}")


def send_ok(endpoint: SerialEndpoint, command: str, timeout_sec: float, context: str) -> CommandResult:
    result = endpoint.send_command(command, timeout_sec)
    if not result.ok:
        raise ValidationRuntimeError(f"{endpoint.config.label} {context} failed: {result.line}")
    return result


def set_result_cell(cell, passed: bool) -> None:
    cell.fill = PASS_FILL if passed else FAIL_FILL
    cell.font = PASS_FONT if passed else FAIL_FONT
    cell.alignment = CELL_ALIGNMENT


def style_header_row(ws, row_index: int) -> None:
    for cell in ws[row_index]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGNMENT


def autosize_sheet(ws) -> None:
    for column in ws.columns:
        values = [str(cell.value) for cell in column if cell.value is not None]
        width = max((len(value) for value in values), default=10)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 10), 40)


def create_workbook(
    output_records: list[dict[str, Any]],
    input_records: list[dict[str, Any]],
    interrupt_records: list[dict[str, Any]],
    report_path: Path,
) -> None:
    workbook = Workbook()

    output_ws = workbook.active
    output_ws.title = "output"
    output_headers = [
        "Group",
        "DUT Pin",
        "Tester Pin",
        "Pull",
        "Drive",
        "Tester Read When DUT=1",
        "High Result",
        "Tester Read When DUT=0",
        "Low Result",
        "Overall Result",
        "Detail",
    ]
    output_ws.append(output_headers)
    style_header_row(output_ws, 1)
    for record in output_records:
        output_ws.append(
            [
                record["group"],
                record["dut_pin"],
                record["tester_pin"],
                record["pull"],
                record["drive"],
                record["high_observation"],
                "PASS" if record["high_pass"] else "FAIL",
                record["low_observation"],
                "PASS" if record["low_pass"] else "FAIL",
                record["result"],
                record["detail"],
            ]
        )
        row = output_ws.max_row
        set_result_cell(output_ws.cell(row=row, column=7), record["high_pass"])
        set_result_cell(output_ws.cell(row=row, column=9), record["low_pass"])
        set_result_cell(output_ws.cell(row=row, column=10), record["result"] == "PASS")
        if not record["high_pass"]:
            output_ws.cell(row=row, column=6).fill = FAIL_FILL
        if not record["low_pass"]:
            output_ws.cell(row=row, column=8).fill = FAIL_FILL
        output_ws.cell(row=row, column=11).alignment = CELL_ALIGNMENT

    input_ws = workbook.create_sheet("input")
    input_headers = [
        "Group",
        "DUT Pin",
        "Tester Pin",
        "Pull",
        "DUT Read When Tester=1",
        "High Result",
        "DUT Read When Tester=0",
        "Low Result",
        "Overall Result",
        "Detail",
    ]
    input_ws.append(input_headers)
    style_header_row(input_ws, 1)
    for record in input_records:
        input_ws.append(
            [
                record["group"],
                record["dut_pin"],
                record["tester_pin"],
                record["pull"],
                record["high_observation"],
                "PASS" if record["high_pass"] else "FAIL",
                record["low_observation"],
                "PASS" if record["low_pass"] else "FAIL",
                record["result"],
                record["detail"],
            ]
        )
        row = input_ws.max_row
        set_result_cell(input_ws.cell(row=row, column=6), record["high_pass"])
        set_result_cell(input_ws.cell(row=row, column=8), record["low_pass"])
        set_result_cell(input_ws.cell(row=row, column=9), record["result"] == "PASS")
        if not record["high_pass"]:
            input_ws.cell(row=row, column=5).fill = FAIL_FILL
        if not record["low_pass"]:
            input_ws.cell(row=row, column=7).fill = FAIL_FILL
        input_ws.cell(row=row, column=10).alignment = CELL_ALIGNMENT

    interrupt_ws = workbook.create_sheet("interrupt")
    interrupt_headers = [
        "Group",
        "DUT Pin",
        "Tester Pin",
        "Pull",
        "Trigger",
        "Idle Level",
        "Active Level",
        "Expected Delta",
        "Retry Count",
        "DUT Response",
        "Result",
        "Detail",
    ]
    interrupt_ws.append(interrupt_headers)
    style_header_row(interrupt_ws, 1)
    for record in interrupt_records:
        interrupt_ws.append(
            [
                record["group"],
                record["dut_pin"],
                record["tester_pin"],
                record["pull"],
                record["trigger"],
                record["idle_level"],
                record["active_level"],
                record["expected_delta"],
                record.get("retry_count", 0),
                record["response"],
                record["result"],
                record["detail"],
            ]
        )
        row = interrupt_ws.max_row
        passed = record["result"] == "PASS"
        set_result_cell(interrupt_ws.cell(row=row, column=11), passed)
        if not passed:
            interrupt_ws.cell(row=row, column=10).fill = FAIL_FILL
        interrupt_ws.cell(row=row, column=12).alignment = CELL_ALIGNMENT

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        autosize_sheet(ws)

    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)


def emit_summary(
    output_records: list[dict[str, Any]],
    input_records: list[dict[str, Any]],
    interrupt_records: list[dict[str, Any]],
) -> dict[str, Any]:
    all_records = output_records + input_records + interrupt_records
    passed = sum(1 for record in all_records if record["result"] == "PASS")
    failed = sum(1 for record in all_records if record["result"] != "PASS")
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "totals": {
            "pass": passed,
            "fail": failed,
            "records": len(all_records),
        },
        "output": output_records,
        "input": input_records,
        "interrupt": interrupt_records,
    }


def run_output_validation(
    dut: SerialEndpoint,
    tester: SerialEndpoint,
    groups: list[PinGroup],
    scenario: dict[str, Any],
    log_fp,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if not scenario or not bool_enabled(scenario.get("enabled"), default=True):
        log_progress(log_fp, "output", "skipped")
        return records

    settle_sec = scenario_timeout_ms(scenario, "settle_ms", 20) / 1000.0
    drive = scenario.get("dut_drive", 0)
    if not isinstance(drive, int) or drive not in (0, 1, 2, 3):
        raise SystemExit("scenarios.output.dut_drive must be one of 0, 1, 2, 3")

    log_progress(log_fp, "output", f"start groups={len(groups)} settle_ms={int(settle_sec * 1000)} drive={drive}")
    for group in groups:
        dut_csv = pins_to_csv(group.dut_pins)
        tester_csv = pins_to_csv(group.tester_pins)
        log_progress(log_fp, "output", f"group={group.name} pins={len(group.dut_pins)} configure tester=in")
        send_ok(tester, f"gpiov_config_many {tester_csv} in", 3.0, "config tester input group")

        for pull in normalize_pull_modes(scenario.get("dut_pull", "float")):
            log_progress(log_fp, "output", f"group={group.name} pull={pull} configure dut=out")
            send_ok(
                dut,
                f"gpiov_config_many {dut_csv} out pull={pull} drive={drive} init=0",
                3.0,
                "config dut output group",
            )
            time.sleep(settle_sec)

            total_pins = len(group.dut_pins)
            for pin_index, (dut_pin, tester_pin) in enumerate(zip(group.dut_pins, group.tester_pins), start=1):
                log_progress(
                    log_fp,
                    "output",
                    f"group={group.name} pull={pull} pin={pin_index}/{total_pins} dut={dut_pin} tester={tester_pin} start",
                )
                high_pass = False
                low_pass = False
                high_observation = ""
                low_observation = ""
                detail_parts: list[str] = []

                try:
                    send_ok(
                        dut,
                        f"gpiov_write_pattern {dut_csv} {dut_pin} 1 0",
                        3.0,
                        f"drive dut pin {dut_pin} high",
                    )
                    time.sleep(settle_sec)
                    high_result = tester.send_command(
                        f"gpiov_read_expect {tester_csv} {tester_pin} 1 0",
                        3.0,
                    )
                    high_pass = high_result.ok
                    high_observation = result_excerpt(high_result)
                    if not high_result.ok:
                        detail_parts.append(f"high: {high_result.line}")
                except Exception as exc:  # pragma: no cover - hardware path
                    high_observation = str(exc)
                    detail_parts.append(f"high: {exc}")

                try:
                    send_ok(
                        dut,
                        f"gpiov_write_pattern {dut_csv} {dut_pin} 0 1",
                        3.0,
                        f"drive dut pin {dut_pin} low",
                    )
                    time.sleep(settle_sec)
                    low_result = tester.send_command(
                        f"gpiov_read_expect {tester_csv} {tester_pin} 0 1",
                        3.0,
                    )
                    low_pass = low_result.ok
                    low_observation = result_excerpt(low_result)
                    if not low_result.ok:
                        detail_parts.append(f"low: {low_result.line}")
                except Exception as exc:  # pragma: no cover - hardware path
                    low_observation = str(exc)
                    detail_parts.append(f"low: {exc}")

                records.append(
                    {
                        "group": group.name,
                        "dut_pin": dut_pin,
                        "tester_pin": tester_pin,
                        "pull": pull,
                        "drive": drive,
                        "high_observation": high_observation,
                        "high_pass": high_pass,
                        "low_observation": low_observation,
                        "low_pass": low_pass,
                        "result": "PASS" if (high_pass and low_pass) else "FAIL",
                        "detail": " | ".join(detail_parts),
                    }
                )
                log_progress(
                    log_fp,
                    "output",
                    "group={group} pull={pull} pin={idx}/{total} dut={dut} tester={tester} result={result} "
                    "high={high_pass} low={low_pass} high_obs={high_obs} low_obs={low_obs}".format(
                        group=group.name,
                        pull=pull,
                        idx=pin_index,
                        total=total_pins,
                        dut=dut_pin,
                        tester=tester_pin,
                        result=records[-1]["result"],
                        high_pass="PASS" if high_pass else "FAIL",
                        low_pass="PASS" if low_pass else "FAIL",
                        high_obs=shorten_text(high_observation),
                        low_obs=shorten_text(low_observation),
                    ),
                )

    log_stage_summary(log_fp, "output", records)
    return records


def run_input_validation(
    dut: SerialEndpoint,
    tester: SerialEndpoint,
    groups: list[PinGroup],
    scenario: dict[str, Any],
    log_fp,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if not scenario or not bool_enabled(scenario.get("enabled"), default=True):
        log_progress(log_fp, "input", "skipped")
        return records

    settle_sec = scenario_timeout_ms(scenario, "settle_ms", 20) / 1000.0

    log_progress(log_fp, "input", f"start groups={len(groups)} settle_ms={int(settle_sec * 1000)}")
    for group in groups:
        dut_csv = pins_to_csv(group.dut_pins)
        tester_csv = pins_to_csv(group.tester_pins)
        log_progress(log_fp, "input", f"group={group.name} pins={len(group.dut_pins)} configure tester=out")
        send_ok(tester, f"gpiov_config_many {tester_csv} out init=0", 3.0, "config tester output group")

        for pull in normalize_pull_modes(scenario.get("dut_pull", "float")):
            log_progress(log_fp, "input", f"group={group.name} pull={pull} configure dut=in")
            send_ok(
                dut,
                f"gpiov_config_many {dut_csv} in pull={pull}",
                3.0,
                "config dut input group",
            )
            time.sleep(settle_sec)

            total_pins = len(group.dut_pins)
            for pin_index, (dut_pin, tester_pin) in enumerate(zip(group.dut_pins, group.tester_pins), start=1):
                log_progress(
                    log_fp,
                    "input",
                    f"group={group.name} pull={pull} pin={pin_index}/{total_pins} dut={dut_pin} tester={tester_pin} start",
                )
                high_pass = False
                low_pass = False
                high_observation = ""
                low_observation = ""
                detail_parts: list[str] = []

                try:
                    send_ok(
                        tester,
                        f"gpiov_write_pattern {tester_csv} {tester_pin} 1 0",
                        3.0,
                        f"drive tester pin {tester_pin} high",
                    )
                    time.sleep(settle_sec)
                    high_result = dut.send_command(
                        f"gpiov_read_expect {dut_csv} {dut_pin} 1 0",
                        3.0,
                    )
                    high_pass = high_result.ok
                    high_observation = result_excerpt(high_result)
                    if not high_result.ok:
                        detail_parts.append(f"high: {high_result.line}")
                except Exception as exc:  # pragma: no cover - hardware path
                    high_observation = str(exc)
                    detail_parts.append(f"high: {exc}")

                try:
                    send_ok(
                        tester,
                        f"gpiov_write_pattern {tester_csv} {tester_pin} 0 1",
                        3.0,
                        f"drive tester pin {tester_pin} low",
                    )
                    time.sleep(settle_sec)
                    low_result = dut.send_command(
                        f"gpiov_read_expect {dut_csv} {dut_pin} 0 1",
                        3.0,
                    )
                    low_pass = low_result.ok
                    low_observation = result_excerpt(low_result)
                    if not low_result.ok:
                        detail_parts.append(f"low: {low_result.line}")
                except Exception as exc:  # pragma: no cover - hardware path
                    low_observation = str(exc)
                    detail_parts.append(f"low: {exc}")

                records.append(
                    {
                        "group": group.name,
                        "dut_pin": dut_pin,
                        "tester_pin": tester_pin,
                        "pull": pull,
                        "high_observation": high_observation,
                        "high_pass": high_pass,
                        "low_observation": low_observation,
                        "low_pass": low_pass,
                        "result": "PASS" if (high_pass and low_pass) else "FAIL",
                        "detail": " | ".join(detail_parts),
                    }
                )
                log_progress(
                    log_fp,
                    "input",
                    "group={group} pull={pull} pin={idx}/{total} dut={dut} tester={tester} result={result} "
                    "high={high_pass} low={low_pass} high_obs={high_obs} low_obs={low_obs}".format(
                        group=group.name,
                        pull=pull,
                        idx=pin_index,
                        total=total_pins,
                        dut=dut_pin,
                        tester=tester_pin,
                        result=records[-1]["result"],
                        high_pass="PASS" if high_pass else "FAIL",
                        low_pass="PASS" if low_pass else "FAIL",
                        high_obs=shorten_text(high_observation),
                        low_obs=shorten_text(low_observation),
                    ),
                )

    log_stage_summary(log_fp, "input", records)
    return records


def run_interrupt_validation(
    dut: SerialEndpoint,
    tester: SerialEndpoint,
    groups: list[PinGroup],
    scenario: dict[str, Any],
    log_fp,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if not scenario or not bool_enabled(scenario.get("enabled"), default=True):
        log_progress(log_fp, "interrupt", "skipped")
        return records

    settle_sec = scenario_timeout_ms(scenario, "settle_ms", 20) / 1000.0
    restore_sec = scenario_timeout_ms(scenario, "restore_settle_ms", 20) / 1000.0
    default_level_hold_ms = scenario_timeout_ms(scenario, "active_hold_ms", int(restore_sec * 1000))
    default_timeout_ms = scenario_timeout_ms(scenario, "timeout_ms", 150)
    irq_read_chunk_size = scenario_timeout_ms(scenario, "irq_read_chunk_size", DEFAULT_IRQ_GET_MANY_CHUNK_SIZE)
    irq_read_retry_max = scenario_non_negative_int(scenario, "irq_read_retry_max", DEFAULT_IRQ_READ_RETRY_MAX)
    irq_read_retry_delay_sec = (
        scenario_non_negative_int(scenario, "irq_read_retry_delay_ms", DEFAULT_IRQ_READ_RETRY_DELAY_MS) / 1000.0
    )
    default_expected_delta = scenario.get("expected_delta", 1)
    if not isinstance(default_expected_delta, int) or default_expected_delta <= 0:
        raise SystemExit("scenarios.interrupt.expected_delta must be a positive integer")

    triggers = parse_interrupt_triggers(scenario)

    log_progress(
        log_fp,
        "interrupt",
        "start groups={groups} settle_ms={settle_ms} restore_ms={restore_ms} triggers={triggers} "
        "irq_read_chunk_size={chunk_size} irq_read_retry_max={retry_max} irq_read_retry_delay_ms={retry_delay_ms}".format(
            groups=len(groups),
            settle_ms=int(settle_sec * 1000),
            restore_ms=int(restore_sec * 1000),
            triggers=len(triggers),
            chunk_size=irq_read_chunk_size,
            retry_max=irq_read_retry_max,
            retry_delay_ms=int(irq_read_retry_delay_sec * 1000),
        ),
    )
    with ThreadPoolExecutor(max_workers=1) as executor:
        for group in groups:
            dut_csv = pins_to_csv(group.dut_pins)
            tester_csv = pins_to_csv(group.tester_pins)

            for pull in normalize_pull_modes(scenario.get("dut_pull", "float")):
                for trigger_cfg in triggers:
                    trigger = trigger_cfg["trigger"]
                    level_trigger = is_level_interrupt_trigger(trigger)
                    idle_level = trigger_cfg["idle_level"]
                    active_level = trigger_cfg["active_level"]
                    expected_delta = trigger_cfg["expected_delta"]
                    timeout_ms = trigger_cfg["timeout_ms"] or default_timeout_ms
                    if trigger_cfg["active_hold_ms"] is not None:
                        active_hold_ms = trigger_cfg["active_hold_ms"]
                    elif level_trigger:
                        active_hold_ms = default_level_hold_ms
                    else:
                        active_hold_ms = int(restore_sec * 1000)
                    active_hold_sec = active_hold_ms / 1000.0
                    if expected_delta is None:
                        expected_delta = 2 if trigger == "sync_both" else default_expected_delta

                    log_progress(
                        log_fp,
                        "interrupt",
                        "group={group} pull={pull} trigger={trigger} configure expected_delta={expected_delta} "
                        "timeout_ms={timeout_ms} active_hold_ms={active_hold_ms}".format(
                            group=group.name,
                            pull=pull,
                            trigger=trigger_cfg["name"],
                            expected_delta=expected_delta,
                            timeout_ms=timeout_ms,
                            active_hold_ms=active_hold_ms,
                        ),
                    )
                    # Exit the previous IRQ mode before changing tester idle levels.
                    # This prevents level-triggered IRQ storms during trigger-to-trigger transitions.
                    send_ok(
                        dut,
                        f"gpiov_config_many {dut_csv} in pull={pull}",
                        3.0,
                        "quiesce dut irq group before reconfigure",
                    )
                    time.sleep(settle_sec)
                    # Drive tester pins to the non-triggering idle level before enabling DUT IRQ mode.
                    # This avoids immediate interrupt storms for low/high level triggers.
                    send_ok(
                        tester,
                        f"gpiov_config_many {tester_csv} out init={idle_level}",
                        3.0,
                        f"config tester output group idle={idle_level}",
                    )
                    time.sleep(settle_sec)
                    send_ok(
                        dut,
                        f"gpiov_config_many {dut_csv} irq pull={pull} trig={trigger}",
                        3.0,
                        f"config dut irq group trigger={trigger}",
                    )
                    time.sleep(settle_sec)

                    total_pins = len(group.dut_pins)
                    for pin_index, (dut_pin, tester_pin) in enumerate(zip(group.dut_pins, group.tester_pins), start=1):
                        log_progress(
                            log_fp,
                            "interrupt",
                            "group={group} pull={pull} trigger={trigger} pin={idx}/{total} dut={dut} tester={tester} "
                            "idle={idle} active={active} active_hold_ms={hold_ms} start".format(
                                group=group.name,
                                pull=pull,
                                trigger=trigger_cfg["name"],
                                idx=pin_index,
                                total=total_pins,
                                dut=dut_pin,
                                tester=tester_pin,
                                idle=idle_level,
                                active=active_level,
                                hold_ms=active_hold_ms,
                            ),
                        )
                        response_text = ""
                        detail_parts: list[str] = []
                        retry_count = 0
                        pin_log_prefix = (
                            "group={group} pull={pull} trigger={trigger} pin={idx}/{total} dut={dut} tester={tester}".format(
                                group=group.name,
                                pull=pull,
                                trigger=trigger_cfg["name"],
                                idx=pin_index,
                                total=total_pins,
                                dut=dut_pin,
                                tester=tester_pin,
                            )
                        )

                        try:
                            send_ok(
                                tester,
                                f"gpiov_write_pattern {tester_csv} {tester_pin} {idle_level} {idle_level}",
                                3.0,
                                f"restore tester group to idle for pin {tester_pin}",
                            )
                            time.sleep(restore_sec)
                            send_ok(dut, f"gpiov_irq_clear_many {dut_csv}", 3.0, "clear dut irq counters")
                            time.sleep(settle_sec)
                            if level_trigger:
                                send_ok(
                                    tester,
                                    f"gpiov_write_pattern {tester_csv} {tester_pin} {active_level} {idle_level}",
                                    3.0,
                                    f"apply level interrupt stimulus on tester pin {tester_pin}",
                                )
                                time.sleep(active_hold_sec)
                                send_ok(
                                    tester,
                                    f"gpiov_write_pattern {tester_csv} {tester_pin} {idle_level} {idle_level}",
                                    3.0,
                                    f"restore tester pin {tester_pin} after level interrupt stimulus",
                                )
                                time.sleep(restore_sec)
                                irq_counts, read_errors, retry_count = read_irq_counts_chunked(
                                    dut,
                                    group.dut_pins,
                                    chunk_size=irq_read_chunk_size,
                                    retry_max=irq_read_retry_max,
                                    retry_delay_sec=irq_read_retry_delay_sec,
                                    log_fp=log_fp,
                                    log_prefix=pin_log_prefix,
                                )
                                response_text, validation_details = evaluate_level_interrupt_result(
                                    group, irq_counts, dut_pin, expected_delta
                                )
                                detail_parts.extend(read_errors)
                                detail_parts.extend(validation_details)
                            else:
                                future = executor.submit(
                                    dut.send_command,
                                    f"gpiov_irq_expect {dut_csv} {dut_pin} {expected_delta} {timeout_ms}",
                                    (timeout_ms / 1000.0) + 1.0,
                                )
                                time.sleep(settle_sec)
                                send_ok(
                                    tester,
                                    f"gpiov_write_pattern {tester_csv} {tester_pin} {active_level} {idle_level}",
                                    3.0,
                                    f"apply interrupt stimulus on tester pin {tester_pin}",
                                )
                                time.sleep(active_hold_sec)
                                send_ok(
                                    tester,
                                    f"gpiov_write_pattern {tester_csv} {tester_pin} {idle_level} {idle_level}",
                                    3.0,
                                    f"restore tester pin {tester_pin} after interrupt stimulus",
                                )
                                time.sleep(restore_sec)
                                irq_result = future.result()
                                response_text = result_excerpt(irq_result)
                                if not irq_result.ok:
                                    detail_parts.append(irq_result.line)
                        except Exception as exc:  # pragma: no cover - hardware path
                            response_text = str(exc)
                            detail_parts.append(str(exc))

                        records.append(
                            {
                                "group": group.name,
                                "dut_pin": dut_pin,
                                "tester_pin": tester_pin,
                                "pull": pull,
                                "trigger": trigger_cfg["name"],
                                "idle_level": idle_level,
                                "active_level": active_level,
                                "expected_delta": expected_delta,
                                "retry_count": retry_count,
                                "response": response_text,
                                "result": "PASS" if not detail_parts else "FAIL",
                                "detail": " | ".join(detail_parts),
                            }
                        )
                        log_progress(
                            log_fp,
                            "interrupt",
                            "group={group} pull={pull} trigger={trigger} pin={idx}/{total} dut={dut} tester={tester} "
                            "result={result} retry_count={retry_count} response={response}".format(
                                group=group.name,
                                pull=pull,
                                trigger=trigger_cfg["name"],
                                idx=pin_index,
                                total=total_pins,
                                dut=dut_pin,
                                tester=tester_pin,
                                result=records[-1]["result"],
                                retry_count=records[-1]["retry_count"],
                                response=shorten_text(response_text),
                            ),
                        )

    log_stage_summary(log_fp, "interrupt", records)
    log_progress(
        log_fp,
        "interrupt",
        f"retry_summary total_retry_count={sum(record.get('retry_count', 0) for record in records)}",
    )
    return records


def run_logged_command(command: list[str], cwd: Path, log_fp, stage_name: str) -> int:
    header = f"\n[{stage_name}] {shlex.join(command)}\n"
    print(header, end="")
    log_fp.write(header)
    log_fp.flush()

    process = subprocess.Popen(
        command,
        cwd=str(cwd),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
    )

    assert process.stdout is not None
    for line in process.stdout:
        print(line, end="")
        log_fp.write(line)
    process.wait()
    log_fp.flush()
    return process.returncode


def resolve_dut_build_command(config: dict[str, Any]) -> list[str] | None:
    targets_cfg = ensure_dict(config.get("targets"), "targets")
    dut_cfg = ensure_dict(targets_cfg.get("dut"), "targets.dut")
    build_cfg = ensure_dict(dut_cfg.get("build"), "targets.dut.build")
    if not ensure_bool(build_cfg.get("enabled"), "targets.dut.build.enabled", True):
        return None

    chip = ensure_str(build_cfg.get("chip"), "targets.dut.build.chip")
    board = ensure_str(build_cfg.get("board"), "targets.dut.build.board")
    make_args = ensure_str_list(build_cfg.get("make_args"), "targets.dut.build.make_args")
    return ["make", f"CHIP={chip}", f"BOARD={board}", *make_args]


def resolve_dut_flash_command(config: dict[str, Any]) -> list[str] | None:
    targets_cfg = ensure_dict(config.get("targets"), "targets")
    dut_cfg = ensure_dict(targets_cfg.get("dut"), "targets.dut")
    flash_cfg = ensure_dict(dut_cfg.get("flash"), "targets.dut.flash")
    if not ensure_bool(flash_cfg.get("enabled"), "targets.dut.flash.enabled", True):
        return None

    build_cfg = ensure_dict(dut_cfg.get("build"), "targets.dut.build")
    serial_cfg = ensure_dict(config.get("serial"), "serial")
    serial_dut_cfg = ensure_dict(serial_cfg.get("dut"), "serial.dut")

    chip = ensure_str(flash_cfg.get("chip", build_cfg.get("chip")), "targets.dut.flash.chip")
    port = ensure_str(flash_cfg.get("port", serial_dut_cfg.get("port")), "targets.dut.flash.port")
    make_args = ensure_str_list(flash_cfg.get("make_args"), "targets.dut.flash.make_args")
    return ["make", "flash", f"CHIP={chip}", f"COMX={port}", *make_args]


def dut_clean_before_build(config: dict[str, Any]) -> bool:
    targets_cfg = ensure_dict(config.get("targets"), "targets")
    dut_cfg = ensure_dict(targets_cfg.get("dut"), "targets.dut")
    build_cfg = ensure_dict(dut_cfg.get("build"), "targets.dut.build")
    return ensure_bool(build_cfg.get("clean_before_build"), "targets.dut.build.clean_before_build", True)


def main() -> int:
    args = parse_args()
    case_dir = Path(__file__).resolve().parent
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = (case_dir / config_path).resolve()
    config = load_json(config_path)

    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
    output_dir = Path(args.output_dir) if args.output_dir else case_dir / "validation_runs" / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)

    report_name = str(config.get("report_name", "gpio_validation_report.xlsx"))
    if not report_name.endswith(".xlsx"):
        report_name += ".xlsx"

    wrapper_log_path = output_dir / "run_full_validation.log"
    report_path = output_dir / report_name
    serial_log_path = output_dir / "gpio_ci_test.log"
    summary_path = output_dir / "gpio_validation_results.json"

    with wrapper_log_path.open("w", encoding="utf-8") as wrapper_log_fp:
        wrapper_log_fp.write(f"config={config_path}\n")
        wrapper_log_fp.write(f"case_dir={case_dir}\n")
        wrapper_log_fp.write(f"output_dir={output_dir}\n")
        wrapper_log_fp.flush()

        build_command = resolve_dut_build_command(config)
        if build_command is not None and not args.skip_build:
            if dut_clean_before_build(config):
                return_code = run_logged_command(["make", "clean"], case_dir, wrapper_log_fp, "dut.clean")
                if return_code != 0:
                    print(f"dut.clean failed, see {wrapper_log_path}")
                    return return_code

            return_code = run_logged_command(build_command, case_dir, wrapper_log_fp, "dut.build")
            if return_code != 0:
                print(f"dut.build failed, see {wrapper_log_path}")
                return return_code

        flash_command = resolve_dut_flash_command(config)
        if flash_command is not None and not args.skip_flash:
            return_code = run_logged_command(flash_command, case_dir, wrapper_log_fp, "dut.flash")
            if return_code != 0:
                print(f"dut.flash failed, see {wrapper_log_path}")
                return return_code

        if args.skip_test:
            print(f"wrapper_log={wrapper_log_path}")
            print(f"output_dir={output_dir}")
            return 0

        serial_cfg = ensure_dict(config.get("serial"), "serial")
        dut_config = parse_serial_config(serial_cfg.get("dut"), "dut")
        tester_config = parse_serial_config(serial_cfg.get("tester"), "tester")
        groups = parse_pin_groups(config.get("pin_groups"))
        scenarios = ensure_dict(config.get("scenarios"), "scenarios")
        startup_drain_sec = scenario_timeout_ms(config, "startup_drain_ms", 500) / 1000.0

        artifact_logger = ArtifactLogger(serial_log_path)
        artifact_logger.log("HOST", f"CONFIG {config_path}")
        try:
            with SerialEndpoint(dut_config, artifact_logger, startup_drain_sec) as dut, SerialEndpoint(
                tester_config, artifact_logger, startup_drain_sec
            ) as tester:
                dut_ping = dut.ping()
                tester_ping = tester.ping()
                artifact_logger.log("HOST", f"DUT_PING {dut_ping.line}")
                artifact_logger.log("HOST", f"TESTER_PING {tester_ping.line}")

                output_records = run_output_validation(
                    dut, tester, groups, scenarios.get("output", {}), wrapper_log_fp
                )
                input_records = run_input_validation(
                    dut, tester, groups, scenarios.get("input", {}), wrapper_log_fp
                )
                interrupt_records = run_interrupt_validation(
                    dut, tester, groups, scenarios.get("interrupt", {}), wrapper_log_fp
                )

            create_workbook(output_records, input_records, interrupt_records, report_path)
            summary = emit_summary(output_records, input_records, interrupt_records)
            summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        finally:
            artifact_logger.close()

    print(f"wrapper_log={wrapper_log_path}")
    print(f"report={report_path}")
    print(f"log={serial_log_path}")
    print(f"summary={summary_path}")
    print(
        "result="
        + ("PASS" if summary["totals"]["fail"] == 0 else "FAIL")
        + f" pass={summary['totals']['pass']} fail={summary['totals']['fail']}"
    )
    return 0 if summary["totals"]["fail"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
